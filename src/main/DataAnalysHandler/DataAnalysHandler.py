import random
import matplotlib.pyplot as plt
import numpy as np
import time
import os
import statistics
import pandas
import openpyxl
import json
import src.main.model_mapping.model_mapping as mapping

from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from loguru import logger

class DataAnalysHandler:
    
    def __init__(self, input_file_path, output_file_path, model_base_path = "D:/pyproj/st/src/main/model_mapping/models.xlsx") -> None:
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.model_mapping = mapping.ModelMapping(model_base_path)
        
    def transfer_xlsx_to_json(self):
        df = pandas.read_excel(self.input_file_path, sheet_name="测速数据收集")
        data = {}
        # 表格列有：环境	触发时间	云端（国内/海外）	云端环境(生产/测试）	版本（编号）	型号	系统	上行	下行
        # 需要将数据转换成：
        # data = {
        #     "环境1":{
        #         "standard_uprate": 123,
        #         "standard_downrate": 123,
        #         "云端1":{
        #             "云端环境1":{
        #                 "版本1":{
        #                     "型号1":{
        #                         "system":"xxx",
        #                         "测试数据": [
        #                             {
        #                                 "uprate": 123,
        #                                 "downrate": 123,
        #                                 "time": "2021-01-01 00:00:00"
        #                             }
        #                         ]
        #                     }
        #                 }
        #             }
        #         }
        #     }
        # }
            
        for i in range(2, len(df)):
            # 访问"环境"列i行值
            env = df.loc[i, "环境"]
            cloud = df.loc[i, "云端（国内/海外）"]
            cloud_env = df.loc[i, "环境(生产/测试)"]
            version = df.loc[i, "版本（编号）"]
            model = df.loc[i, "型号"]
            system = df.loc[i, "系统"]
            uprate = df.loc[i, "上行"]
            downrate = df.loc[i, "下行"]
            
            # 判断data中是否有该key
            if env not in data:
                data[env] = {}
            if cloud not in data[env]:
                data[env][cloud] = {}
            if cloud_env not in data[env][cloud]:
                data[env][cloud][cloud_env] = {}
            if version not in data[env][cloud][cloud_env]:
                data[env][cloud][cloud_env][version] = {}
            if model not in data[env][cloud][cloud_env][version]:
                data[env][cloud][cloud_env][version][model] = {}
                data[env][cloud][cloud_env][version][model]["system"] = system
                data[env][cloud][cloud_env][version][model]["测试数据"] = []
            data[env][cloud][cloud_env][version][model]["测试数据"].append({
                "uprate": uprate,
                "downrate": downrate,
                "time": df.loc[i, "触发时间"]
            })
        
        return data
    
    def analyse_data(self, data):
        # 计算每个环境-云端-云端环境-版本-型号的:
        # 1. 上行平均值
        # 2. 下行平均值
        # 3. 上行波动值: 使用标准差计算
        # 4. 下行波动值: 使用标准差计算
        # 5. 上行波动范围: 最大值-最小值
        # 6. 下行波动范围: 最大值-最小值
        # 输出数据:
        # output = [
        #     {
        #         "环境": "环境1",
        #         "云端": "云端1",
        #         "版本": "版本1",
        #         "设备型号": "型号1_系统1",
        #         "上行平均值": 123,
        #         "下行平均值": 123,
        #         "上行波动值": 123,
        #         "下行波动值": 123,
        #         "上行波动范围": 123,
        #         "下行波动范围": 123,
        #         "上行分布图": "path/to/figure1.png",
        #         "下行分布图": "path/to/figure2.png",
        #         "样本数量": 123
        #     },
        #     {...}
        # ]
        # 先遍历env中的版本，判断是否有“信通院全球网测”，如果有, 将每种机型下的上下行速率分别合并成数组，并从大到小排列, 生成标准数据
        for env in data:
            logger.info(data[env].keys())
            if "非睿易应用" in data[env].keys() and "信通院全球网测" in data[env]["非睿易应用"]["非睿易应用"]:
                up_result = []
                down_result = []
                for device in data[env]["非睿易应用"]["非睿易应用"]["信通院全球网测"]:
                    devices = data[env]["非睿易应用"]["非睿易应用"]["信通院全球网测"]
                    up_result.extend([i["uprate"] for i in devices[device]["测试数据"]])
                    down_result.extend([i["downrate"] for i in devices[device]["测试数据"]])
                # 将两个列表分别从大到小排序
                up_result.sort(reverse=True)
                down_result.sort(reverse=True)
                # 取标准值：如果最大值与中位数偏差较大的情况下取较小的下一个， 直到数值和中位数偏差40%以内
                up_middle = statistics.median(up_result)
                down_middle = statistics.median(down_result)
                up_standard = up_result[[i for i in range(len(up_result)) if (abs(up_result[i] - up_middle) / up_middle < 0.4)][0]]
                down_standard = down_result[[i for i in range(len(down_result)) if (abs(down_result[i] - down_middle) / down_middle < 0.4)][0]]
                data[env]["up_standard"] = up_standard
                data[env]["up_standard_list"] = up_result
                data[env]["down_standard"] = down_standard
                data[env]["down_standard_list"] = down_result
                logger.info("up_standard: {}, down_standard: {}", up_standard, down_standard)
                
        output = []
        # print(data)
        for env in data:
            for cloud in data[env]:
                # print(cloud)
                if cloud == "up_standard" or cloud == "down_standard" or cloud == "up_standard_list" or cloud == "down_standard_list":
                    continue
                for cloud_env in data[env][cloud]:
                    # print(data[env][cloud][cloud_env])
                    for version in data[env][cloud][cloud_env]:
                        # print(version)
                        for model in data[env][cloud][cloud_env][version]:
                            uprates = []
                            downrates = []
                            for test_data in data[env][cloud][cloud_env][version][model]["测试数据"]:
                                uprates.append(test_data["uprate"])
                                downrates.append(test_data["downrate"])
                            # 计算样本平均值
                            uprate_mean = statistics.mean(uprates)
                            downrate_mean = statistics.mean(downrates)
                            # 计算总体标准差
                            uprates_std = statistics.stdev(uprates) if len(uprates) > 1 else 0
                            downrates_std = statistics.stdev(downrates) if len(downrates) > 1 else 0
                            # 计算波动范围
                            uprate_range = max(uprates) - min(uprates)
                            downrate_range = max(downrates) - min(downrates)
                            # 绘制上下行偏差分布图
                            path = "src/main/DataAnalysHandler/figures/"
                            uprate_file_path = self.generate_comparison_bar_plot(uprates, data[env]["up_standard_list"] if "up_standard_list" in data[env] else [],data[env]["up_standard"] if "up_standard" in data[env] else 0, path, f"{env}_{cloud}_{cloud_env}_{version}_{model}_uprate")
                            downrate_file_path = self.generate_comparison_bar_plot(downrates, data[env]["down_standard_list"] if "down_standard_list" in data[env] else [], data[env]["down_standard"] if "down_standard" in data[env] else 0, path, f"{env}_{cloud}_{cloud_env}_{version}_{model}_downrate")
                            output.append({
                                "环境": env,
                                "云端": cloud,
                                "版本": version,
                                "设备型号": model,
                                "上行平均值": uprate_mean,
                                "下行平均值": downrate_mean,
                                "上行波动值": uprates_std,
                                "下行波动值": downrates_std,
                                "上行波动范围": uprate_range,
                                "下行波动范围": downrate_range,
                                "上行分布图": uprate_file_path,
                                "下行分布图": downrate_file_path,
                                "样本数量": len(uprates),
                                "上行参考值": data[env]["up_standard"] if "up_standard" in data[env] else 0,
                                "下行参考值": data[env]["down_standard"] if "down_standard" in data[env] else 0
                            })
        return output
    
    def save_data_to_xlsx(self, data, output_file_path):
        # data 格式:
        # data = [
        #     {
        #         "环境": "环境1",
        #         "云端": "云端1",
        #         "版本": "版本1",
        #         "设备型号": "型号1_系统1",
        #         "上行平均值": 123,
        #         "下行平均值": 123,
        #         "上行波动值": 123,
        #         "下行波动值": 123,
        #         "上行波动范围": 123,
        #         "下行波动范围": 123,
        #         "上行分布图": "path/to/figure1.png",
        #         "下行分布图": "path/to/figure2.png",
        #         "样本数量": 123
        #     },
        #     {...}
        # ]
        # 创建一个新的Excel文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # "环境"第一列, 云端第二列, 版本第三列, 
        # 上行平均值第一行, 上行波动值第二行, 上行波动范围第三行, 上行分布图第四行, 下行同理
        # 写入表头: 环境, 云端, 版本, 详细信息
        sheet["A1"] = "环境"
        sheet["B1"] = "版本"
        sheet["C1"] = "设备型号"
        sheet["D1"] = "详细信息"
        sheet["E1"] = "详细信息"
        unit = 11
        counter = 0
        imgsize = (1280 / 3, 720 / 3)
        sheet.column_dimensions["E"].width = imgsize[0]*0.14
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        for i in data:
            i["设备型号"] = self.model_mapping.mapping(i["设备型号"])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "样本数量", i["样本数量"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "上行参考值", i["上行参考值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "上行平均值", i["上行平均值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "上行标准差", i["上行波动值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "上行波动范围", i["上行波动范围"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "上行相对参考值偏差分布图" if i["上行参考值"] != 0 else "上行相对平均值偏差分布图"])
            img = Image(i["上行分布图"])
            img.width, img.height = imgsize
            sheet.add_image(img, f"E{unit*counter+7}")
            sheet.row_dimensions[unit*counter+7].height = imgsize[1]*0.8

            sheet.append([i["环境"], i["版本"], i["设备型号"], "下行参考值", i["下行参考值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "下行平均值", i["下行平均值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "下行标准差", i["下行波动值"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "下行波动范围", i["下行波动范围"]])
            sheet.append([i["环境"], i["版本"], i["设备型号"], "下行相对参考值偏差分布图" if i["下行参考值"] != 0 else "下行相对平均值偏差分布图"])
            img = Image(i["下行分布图"])
            img.width, img.height = imgsize
            sheet.add_image(img, f"E{unit*counter+12}")
            sheet.row_dimensions[unit*counter+12].height = imgsize[1]*0.8

            counter += 1
        # 从第二行开始, 每9行分别合并1-4列单元格
        for i in range(2, len(data)*unit, unit):
            sheet.merge_cells(f"A{i}:A{i+unit-1}")
            sheet.merge_cells(f"B{i}:B{i+unit-1}")
            sheet.merge_cells(f"C{i}:C{i+unit-1}")
        # 设置第5列数值居左
        for i in range(2, len(data)*unit):
            sheet[f"E{i}"].alignment = Alignment(horizontal='left')
        book_name = output_file_path+"output.xlsx" if output_file_path[-1] == "/" else output_file_path+"/output.xlsx"
        # 保存文件
        workbook.save(book_name)
        return
    
    def execute(self, input_file_path, output_file_path):
        data = self.transfer_xlsx_to_json()
        # 将data写入output.json
        # print(data)
        output = self.analyse_data(data)
        self.save_data_to_xlsx(output, output_file_path)
                    
                            
    def generate_comparison_bar_plot(self, data, sample_data, standard_value, folder_path, fig_name):
        # 计算偏离百分比
        deviation_percent = [(x - standard_value) / standard_value * 100 for x in data]
        sample_deviation_percent = [(x - standard_value) / standard_value * 100 for x in sample_data]

        # 分组范围
        bins = np.arange(-100, 110, 10)

        # 计算每组的百分比
        data_hist, _ = np.histogram(deviation_percent, bins)
        sample_hist, _ = np.histogram(sample_deviation_percent, bins)
        data_percent = data_hist / sum(data_hist) * 100
        sample_percent = sample_hist / sum(sample_hist) * 100

        # 绘制柱状图
        plt.figure(figsize=(12, 6))
        # plt.bar(bins[:-1], data_percent, width=9, label='Sample', alpha=0.7)
        bars1 = plt.bar(bins[:-1], data_percent, width=9, label='Sample', alpha=0.7)
        if len(sample_data) > 0:
            # plt.bar(bins[:-1], -sample_percent, width=9, label='Standard Sample', alpha=0.7)
            bars2 = plt.bar(bins[:-1], -sample_percent, width=9, label='Standard Sample', alpha=0.7)

        plt.axhline(0, color='black', linewidth=0.8)  # 横轴
        plt.xticks(bins)
        plt.xlabel('Deviation Percentage (%)')
        plt.ylabel('Percentage of Data Points (%)')
        plt.title('Comparison of Data and Sample Data Deviation')
        plt.legend()

        # 在柱状图上添加文本
        for bar, percent, count in zip(bars1, data_percent, data_hist):
            height = bar.get_height()
            x_position = bar.get_x() + bar.get_width() / 2.0
            if np.isfinite(height) and np.isfinite(x_position):
                plt.text(x_position, height, f'{percent:.1f}%\n({count})', ha='center', va='bottom' if height < 0 else 'top')

        if len(sample_data) > 0:
            for bar, percent, count in zip(bars2, sample_percent, sample_hist):
                height = bar.get_height()
                x_position = bar.get_x() + bar.get_width() / 2.0
                if np.isfinite(height) and np.isfinite(x_position):
                    plt.text(x_position, height, f'{percent:.1f}%\n({count})', ha='center', va='top' if height < 0 else 'bottom')

        plt.savefig(f"{folder_path}/{fig_name}.png")
        plt.close()
                
        # 获取工作目录
        file_name = f"{fig_name}.png"
        file_path = os.path.join(folder_path, file_name)

        # 返回图片文件的路径
        return os.getcwd().replace("\\", "/") + "/" + file_path if file_path[1] != ":" else file_path


def main():
    input_file_path = "C:/Users/eureka/Downloads/测速数据收集 (4).xlsx"
    output_file_path = "D:/pyproj/st/src/output"
    dah = DataAnalysHandler(input_file_path, output_file_path)
    dah.execute(input_file_path, output_file_path)

if __name__ == "__main__":
    main()