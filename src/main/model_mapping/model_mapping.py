import openpyxl

class ModelMapping:

    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = self.load_workbook()
    
    def load_workbook(self):
        return openpyxl.load_workbook(self.excel_file_path)
    def mapping(self, model):
        # 打开models sheet
        sheet = self.wb['models']
        # 遍历第一列， 将第一列x行值与model比较，找到对应的行
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == model:
                    # 找到model对应的行，将行的第二列值返回
                    return sheet.cell(row=cell.row, column=7).value+' '+sheet.cell(row=cell.row, column=8).value.replace('#', '')
        return model

def main():
    excel_file_path = 'src/main/model_mapping/models.xlsx'
    modelMapping = ModelMapping(excel_file_path)
    print(modelMapping.mapping('BMH-AN10'))
    print(modelMapping.mapping('ALN-AL80'))

if __name__ == '__main__':
    main()